import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

# Rutas de los archivos

archivo_original = r"2022 Fase 2- Depliegue del Universo de Data.xlsx"
archivo_nuevo = r"2023 Fase 2- Depliegue del Universo de Data.xlsx"

# Funciones

def abrir_archivo(ruta):
    """ Abre un archivo de Excel y devuelve el objeto workbook. """
    try:
        return load_workbook(ruta)
    except Exception as e:
        print(f"Error al abrir el archivo {ruta}: {e}")
        return None
    
def aplicar_formato(celda, negrita=False, color=None):
    """ Aplica formato a una celda. """
    if negrita:
        celda.font = Font(bold=True)
    if color:
        celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

def crear_copia_archivo(archivo_origen, archivo_destino):
    """ Crea una copia del archivo. """
    try:
        shutil.copy(archivo_origen, archivo_destino)
        print("Archivo copiado con éxito.")
    except FileNotFoundError:
        print("Error: El archivo original no se encontró en la ruta especificada.")
    except Exception as e:
        print(f"Se produjo un error inesperado al copiar el archivo: {e}")

def crear_copia_pestaña(ruta_archivo):
    """ Procesa el archivo de Excel para crear o verificar una pestaña. """
    try:
        wb = abrir_archivo(ruta_archivo)
        if "Revisado2023" in wb.sheetnames:
            print("La pestaña 'Revisado2023' ya existe.")
        else:
            source = wb["Revisado2022"]
            target = wb.copy_worksheet(source)
            target.title = "Revisado2023"
            wb.save(ruta_archivo)
            print("La pestaña 'Revisado2023' ha sido creada con éxito.")
        wb.close()
    except InvalidFileException:
        print("Error: El archivo de Excel no es válido o está corrupto.")
    except KeyError:
        print("Error: La pestaña 'Revisado2022' no se encuentra en el archivo.")
    except Exception as e:
        print(f"Se produjo un error inesperado al trabajar con el archivo de Excel: {e}")

# Ejecución del programa:

# Copiar archivo original a una nueva ubicación
crear_copia_archivo(archivo_original, archivo_nuevo)

# Copiar pestaña Revisado2022 a una nueva pestaña y renombrarla como Revisado2023
crear_copia_pestaña(archivo_nuevo)

# Rutas para el procesamiento de archivos adicionales
ruta_primer_archivo = archivo_nuevo
ruta_segundo_archivo = r"DATA - Anexo 12 - 2023"

# Abrir y procesar el primer archivo
wb_primer_archivo = abrir_archivo(ruta_primer_archivo)
hoja_primer_archivo = wb_primer_archivo["Revisado2023"]

# Procesar cada archivo en el directorio especificado
for archivo in os.listdir(ruta_segundo_archivo):
    if archivo.endswith(".xlsx"):
        ruta_completa = os.path.join(ruta_segundo_archivo, archivo)
        wb_segundo_archivo = abrir_archivo(ruta_completa)
        hoja_matriz_de_pruebas = wb_segundo_archivo["Matriz de Pruebas "]
        hoja_controles_sin_prueba = wb_segundo_archivo["Controles sin prueba"]
        nombre = archivo[:10]
        proyecto = ''
        
        # Filas en el primer archivo donde la columna "C" es igual a "nombre"
        for fila in hoja_primer_archivo.iter_rows(min_row=2):
            if fila[2].value == nombre:
                proyecto = fila[3].value
                actualizacion_realizada = False
                # Comparar con cada fila de la hoja matriz de pruevas
                for fila_segundo in hoja_matriz_de_pruebas.iter_rows(min_row=5):
                    if fila[5].value == fila_segundo[2].value and fila[6].value == fila_segundo[3].value and fila[10].value == fila_segundo[16].value:
                        actualizacion_realizada = True
                        # Poner en negrita los cambios o pintar de plomo
                        columnas_a_comparar = [(7, 6), (8, 9), (9, 15), (11, 18), (12, 34), (13, 56), (14, 78)]
                        for col_primer_archivo, col_segundo_archivo in columnas_a_comparar:
                            if fila[col_primer_archivo].value != fila_segundo[col_segundo_archivo].value:
                                fila[col_primer_archivo].value = fila_segundo[col_segundo_archivo].value
                                aplicar_formato(fila[col_primer_archivo], negrita=True)
                                fila[0].value = 'c'
                if not actualizacion_realizada:
                    for celda in fila:
                        aplicar_formato(celda, color="757171")

                for fila_segundo in hoja_controles_sin_prueba.iter_rows(min_row=5):
                    if fila[5].value == fila_segundo[2].value and fila[6].value == fila_segundo[3].value and fila[10].value == fila_segundo[16].value:
                        for celda in fila:
                            aplicar_formato(celda, color='FFFFFF')
                        # Comprobar y actualizar las columnas especificadas
                        columnas_a_comparar = [(7, 6), (8, 9), (9, 15), (11, 19), (12, 35), (13, 57), (14, 79)]
                        for col_primer_archivo, col_segundo_archivo in columnas_a_comparar:
                            if fila[col_primer_archivo].value != fila_segundo[col_segundo_archivo].value:
                                fila[col_primer_archivo].value = fila_segundo[col_segundo_archivo].value
                                aplicar_formato(fila[col_primer_archivo], negrita=True)
                                fila[0].value = 's'
        
        # Nueva lógica para insertar filas en el primer archivo si se cumplen ciertas condiciones
        for fila_segundo in hoja_matriz_de_pruebas.iter_rows(min_row=5):
            existe_coincidencia = False
            for fila in hoja_primer_archivo.iter_rows(min_row=2):
                if (fila[5].value == fila_segundo[2].value and 
                    fila[6].value == fila_segundo[3].value and 
                    fila[10].value == fila_segundo[16].value):
                    existe_coincidencia = True
                    break
            if not existe_coincidencia:
                # Insertar una nueva fila al final del primer archivo
                nueva_fila = hoja_primer_archivo.max_row + 1
                hoja_primer_archivo.cell(row=nueva_fila, column=3, value=nombre)
                hoja_primer_archivo.cell(row=nueva_fila, column=4, value=proyecto)
                columnas_a_copiar = [(2, 6), (3, 7), (6, 8), (9, 9), (15, 10), (16, 11), (18, 12), (34, 13), (56, 14), (78, 15)]
                for col_segundo_archivo, col_primer_archivo in columnas_a_copiar:
                    hoja_primer_archivo.cell(row=nueva_fila, column=col_primer_archivo, value=fila_segundo[col_segundo_archivo].value)
                # Aplicar formato a la nueva fila
                for celda in hoja_primer_archivo[nueva_fila]:
                    aplicar_formato(celda, color="66FF33")
                hoja_primer_archivo.cell(row=nueva_fila, column=1, value='c')
        
        # Nueva lógica para insertar filas en el primer archivo desde la pestaña "Controles sin prueba"
        for fila_segundo in hoja_controles_sin_prueba.iter_rows(min_row=5):
            existe_coincidencia = False
            for fila in hoja_primer_archivo.iter_rows(min_row=2):
                if (fila[5].value == fila_segundo[2].value and 
                    fila[6].value == fila_segundo[3].value and 
                    fila[10].value == fila_segundo[16].value):
                    existe_coincidencia = True
                    break
            if not existe_coincidencia:
                # Insertar una nueva fila al final del primer archivo
                nueva_fila = hoja_primer_archivo.max_row + 1
                hoja_primer_archivo.cell(row=nueva_fila, column=3, value=nombre)
                hoja_primer_archivo.cell(row=nueva_fila, column=4, value=proyecto)
                columnas_a_copiar = [(2, 6), (3, 7), (6, 8), (9, 9), (15, 10), (16, 11), (19, 12), (35, 13), (57, 14), (79, 15)]
                for col_segundo_archivo, col_primer_archivo in columnas_a_copiar:
                    hoja_primer_archivo.cell(row=nueva_fila, column=col_primer_archivo, value=fila_segundo[col_segundo_archivo].value)
                # Aplicar formato a la nueva fila
                for celda in hoja_primer_archivo[nueva_fila]:
                    aplicar_formato(celda, color="66FF33")
                hoja_primer_archivo.cell(row=nueva_fila, column=1, value='s')

        wb_segundo_archivo.close()

# Guardar y cerrar el primer archivo
wb_primer_archivo.save(ruta_primer_archivo)
wb_primer_archivo.close()
